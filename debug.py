import birdnest
import os
import openpyxl
import json
from cleantext import clean
def printfull():
    tweet_row = birdnest.t_dump_by_row('text')
    author = birdnest.t_dump_by_row('authorUSN')
    time = birdnest.t_dump_by_row('time')
    for x in range(0, len(tweet_row)):
        print("---------------")
        print(author[x][0])
        print(tweet_row[x][0])

def youtubetoreadable():
    path = os.walk("./memory/youtube")
    pathstr = "./memory/youtube/"
    slots = []
    header = [['videoID', 'name', 'date', 'text', 'start_second', 'end_second']]
    for root, directories, files in path:
        for file in files:
            file1 = open(pathstr+file, 'r')
            Lines = file1.readlines()
            count = 0
            vid_id =file[0:file.find(".txt")]
            vid_id = vid_id[vid_id.find("_")+1:len(vid_id)]
            for line in Lines:
                line = line.strip()
                listsep = line.split(",")
                name = listsep[0].split(":")[0]
                transtext = listsep[0].split(":")[1]
                starttime = listsep[1]
                runtime = listsep[2]
                date = listsep[3]
                currentslot = [vid_id, name, date, transtext, starttime, runtime]
                slots.append(currentslot)
    slots = header + slots
    return(slots)

def convert_youtube_spreadsheet():
    path = "./memory/youtube.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    slots = youtubetoreadable()
    for x in range(0, len(slots)):
        for y in range(0, len(slots[x])):
            cc = sheet.cell(row = x+1, column = y+1)
            cc.value = slots[x][y]
    workbook.save(filename=path)

#whole mess of cleaning that is waiting to be broken by a bad encoding
def emailtoreadable():
    slots = []
    header = [['from', 'subject', 'message']]
    with open('memory/inbox_old.txt', 'r') as file:
        data = file.read().replace('\n', '')
    data = data.split("==INTERRUPT==")
    print(len(data))
    q = {}
    for x in range(0, len(data)):
        data[x]= clean(data[x],fix_unicode=True,to_ascii=False, no_line_breaks=True)
        data[x] = clean(data[x], no_urls=True,replace_with_url="<URL>")
        data[x]=data[x].replace(",", ":")
        splitbycomma = data[x].split(":",)
        from_mark = None
        message_mark = None
        subject_mark = None
        for y in range(0, len(splitbycomma)):
            #print(splitbycomma[y])

            if('from' in splitbycomma[y] and from_mark is None):
                from_mark = y
                #print("from:")
                #print(splitbycomma[y])
            if('subject' in splitbycomma[y] and subject_mark is None):
                subject_mark = y
                #print("subject:")
                #print(splitbycomma[y])
            if('message' in splitbycomma[y] and message_mark is None):
                message_mark = y
                #print("message:")
                #print(splitbycomma[y])

            #print(subject_mark)
            subject = " ".join(splitbycomma[subject_mark:from_mark])
            #print(from_mark)
            froms = " ".join(splitbycomma[from_mark:message_mark])
            #print(message_mark)
            message = " ".join(splitbycomma[message_mark:len(splitbycomma)])


        subject = subject.replace("{", "")
        subject = subject.replace("}", "")
        subject = subject.replace("'", "")
        subject = subject[subject.find('  '):len(subject)]


        froms = froms.replace("{", "")
        froms = froms.replace("}", "")
        froms = froms.replace("'", "")
        froms = froms[froms.find("from")+len("from"):len(froms)]


        message = message.replace("{", "")
        message = message.replace("}", "")
        message = message.replace("'", "")
        message = message[message.find("[")+1:len(message)-1]
        slots.append([froms, subject, message])
    slots = header+slots
    print(len(slots))
    return(slots)
    pass

def convert_email_spreadsheet():
    path = "./memory/inbox.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    slots = emailtoreadable()
    for x in range(0, len(slots)):
        for y in range(0, len(slots[x])):
            cc = sheet.cell(row = x+1, column = y+1)
            cc.value = slots[x][y]
    workbook.save(filename=path)
convert_youtube_spreadsheet()
convert_email_spreadsheet()
